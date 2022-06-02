import openpyxl

book=openpyxl.load_workbook("C:\\Users\\afraj\\Desktop\\Rahulsheety.xlsx")
sheet=book.active
#print(sheet['C2'].value)
#print(sheet.max_row)
#print(sheet.max_column)
#sheet['D1'].value="Rahul"
#print(sheet.cell(row=1, column=4).value)
#for i in range(2, sheet.max_row+1):
    #if sheet.cell(row=i, column=1).value=="Testcase_1":
        #for j in range(2, sheet.max_column+1):
            #print(sheet.cell(row=i, column=j).value)
dict={}
for i in range(2, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value=="Test_case2":
        for j in range(2, sheet.max_column+1):
            dict[sheet.cell(row=1, column=j).value]=sheet.cell(row=i, column=j).value
print(dict)
