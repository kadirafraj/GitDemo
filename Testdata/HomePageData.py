class HomePageData:
    #test_homepagedata=[{"firstname":"Rahul", "email":"Shetty","password":"Rahul@abc","gender":"Male"}, {"firstname":"Anshika", "email":"Shetty","password":"Anshika@abc","gender":"Female"}]
    @staticmethod
    def getTestData():
        dict = {}
        import openpyxl

        book = openpyxl.load_workbook("C:\\Users\\afraj\\Desktop\\Rahulsheety.xlsx")
        sheet = book.active
        for i in range(2, sheet.max_row + 1):
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[dict]
